"""

Google Sheets 存取層。

本模組專為《健身教練管理系統》設計，提供以下工作表的 CRUD：
- Users        ：教練與學員帳號資訊
- Records      ：學員每日飲食記錄
- Weight       ：學員體重記錄
- Training     ：學員訓練記錄（重量訓練/有氧訓練/其他）

這些存取全走 .streamlit/secrets.toml 簡寫，沒有其他設定。
工作表 schema 由 tools/init_sheets.py 明確建立；正式請求只驗證、不修改 schema。

使用範例：
- get_users_rows / append_user / set_user_role
- get_records / append_record
- get_user_goals / update_user_goals
- get_weight_records / append_weight
- get_training_records / append_training
"""

from __future__ import annotations

from datetime import date
from collections.abc import Iterable
import hashlib
import json
import os
import threading
import time
from typing import Any
import uuid

import gspread
from google.oauth2.service_account import Credentials
import streamlit as st

from config.settings import SETTINGS
from domain.history import summarize_weight_measurements
from domain.validation import (
    TEXT_LIMITS,
    bounded_text,
    finite_non_negative,
    meal_type as validate_meal_type,
    positive_number,
    valid_timestamp,
)

CACHE_TTL = SETTINGS.cache_ttl_seconds
FIXED_PRIMARY_COACH_ID = SETTINGS.primary_coach_id
_VERIFIED_WORKSHEET_SCHEMAS: set[tuple[str, str]] = set()
_APPLIED_IMPORT_TOKENS: set[str] = set()
_IMPORT_TOKEN_LOCK = threading.RLock()


def _retry_transient(operation, *, attempts: int = 3):
    """Retry quota/network/server failures; validation errors are never retried."""
    for attempt in range(attempts):
        try:
            return operation()
        except (ValueError, TypeError):
            raise
        except Exception as exc:
            response = getattr(exc, "response", None)
            status = getattr(response, "status_code", None) or getattr(exc, "code", None)
            if status not in {429, 500, 502, 503, 504} or attempt == attempts - 1:
                raise
            time.sleep(0.25 * (2**attempt))

# ---------- Headers ----------

USERS_HEADERS = [
    "user_id",
    "username",
    "name",           # 學員真實姓名
    "password_hash",
    "created_at",
    "bmr",
    "daily_calorie_goal",
    "daily_protein_goal",
    "daily_carb_goal",
    "daily_fat_goal",
    "daily_water_goal",
    "role",
    "weekly_training_goal",
    "record_mode",          # "simple" 或 "full"，學員的飲食記錄模式
    "coach_id",            # 所屬教練 ID；教練帳號與未分配學員留空
    "must_change_password", # 臨時密碼登入後必須先更改密碼
]

RECORDS_HEADERS = [
    "timestamp",
    "user_id",
    "meal_type",
    "food_summary",
    "calories",
    "protein",
    "carb",
    "fat",
    "water_ml",
    "image_url",
    "portion",
]

# 體重記錄工作表
WEIGHT_HEADERS = [
    "timestamp",
    "user_id",
    "weight_kg",
]

# 訓練記錄工作表
TRAINING_HEADERS = [
    "timestamp",
    "user_id",
    "training_types",
    "strength_detail",
    "cardio_detail",
    "other_detail",
]

TRAINING_TYPE_FIELDS = {
    "重量訓練": "strength_detail",
    "有氧訓練": "cardio_detail",
    "其他": "other_detail",
}

# 教練備註工作表
NOTES_HEADERS = [
    "timestamp",       # 備註時間
    "user_id",       # 學員 ID
    "coach_id",      # 教練 ID
    "note",          # 備註內容
]

AUDIT_HEADERS = [
    "timestamp",
    "request_id",
    "actor_id",
    "actor_role",
    "action",
    "target_type",
    "target_id",
    "result",
    "metadata_json",
]

PASSWORD_RESET_HEADERS = [
    "request_id",
    "user_id",
    "requested_at",
    "status",
    "resolved_at",
    "resolved_by",
]

WORKSHEET_SCHEMAS = {
    "Users": USERS_HEADERS,
    "Records": RECORDS_HEADERS,
    "Weight": WEIGHT_HEADERS,
    "Training": TRAINING_HEADERS,
    "Notes": NOTES_HEADERS,
    "AuditLog": AUDIT_HEADERS,
    "PasswordResetRequests": PASSWORD_RESET_HEADERS,
}


# ---------- 內部工具 ----------

def _get_secrets() -> dict[str, Any]:

    if "gcp" not in st.secrets or "SPREADSHEET_ID" not in st.secrets:
        raise EnvironmentError(
            "請先在 .streamlit/secrets.toml 中設定 [gcp] (含 Service Account JSON) 與 SPREADSHEET_ID"
        )
    gcp = dict(st.secrets["gcp"])
    gcp["SPREADSHEET_ID"] = st.secrets["SPREADSHEET_ID"]
    return gcp


@st.cache_resource(show_spinner=False)
def _get_client() -> gspread.Client:
    sec = _get_secrets()
    sa_info = {k: v for k, v in sec.items() if k != "SPREADSHEET_ID"}
    creds = Credentials.from_service_account_info(
        sa_info,
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ],
    )
    return gspread.authorize(creds)


@st.cache_resource(show_spinner=False)
def _get_sheet() -> gspread.Spreadsheet:
    if os.environ.get("PROJECT_PRIME_TESTING") == "1":
        raise RuntimeError("測試環境禁止連線正式 Google Sheet")
    sec = _get_secrets()
    return _get_client().open_by_key(sec["SPREADSHEET_ID"])


def _ensure_worksheet(sh: gspread.Spreadsheet, title: str, headers: list[str]) -> gspread.Worksheet:
    """Validate a worksheet without mutating production data or schema."""
    try:
        ws = sh.worksheet(title)
    except gspread.WorksheetNotFound as exc:
        raise RuntimeError(f"缺少 {title} 工作表，請先執行 tools/init_sheets.py") from exc

    schema_key = (str(getattr(sh, "id", id(sh))), title)
    if schema_key in _VERIFIED_WORKSHEET_SCHEMAS:
        return ws
    existing_headers = ws.row_values(1)
    if existing_headers != headers:
        raise RuntimeError(
            f"{title} 工作表欄位與程式定義不一致，請先執行 tools/init_sheets.py"
        )
    _VERIFIED_WORKSHEET_SCHEMAS.add(schema_key)
    return ws


def initialize_worksheets(*, apply: bool = False) -> list[dict[str, Any]]:
    """Preview or explicitly create/extend known worksheets."""
    sh = _get_sheet()
    report: list[dict[str, Any]] = []
    for title, headers in WORKSHEET_SCHEMAS.items():
        try:
            ws = sh.worksheet(title)
        except gspread.WorksheetNotFound:
            report.append({"worksheet": title, "status": "missing"})
            if apply:
                ws = sh.add_worksheet(title=title, rows=1000, cols=len(headers))
                ws.append_row(headers, value_input_option="RAW")
            continue
        existing = ws.row_values(1)
        if existing == headers:
            report.append({"worksheet": title, "status": "ok"})
        elif existing == headers[: len(existing)]:
            missing = headers[len(existing) :]
            report.append({"worksheet": title, "status": "extend", "headers": missing})
            if apply:
                if ws.col_count < len(headers):
                    ws.add_cols(len(headers) - ws.col_count)
                for col, header in enumerate(missing, start=len(existing) + 1):
                    ws.update_cell(1, col, header)
        else:
            report.append({"worksheet": title, "status": "mismatch"})
    if apply:
        _VERIFIED_WORKSHEET_SCHEMAS.clear()
    return report


def _rows_to_dicts(ws: gspread.Worksheet, headers: list[str]) -> list[dict[str, Any]]:
    """將工作表列（第一列為 header）轉為 dict 列表。"""
    values = ws.get_all_values()
    if not values:
        return []
    out: list[dict[str, Any]] = []
    for row in values[1:]:
        record = {}
        for i, key in enumerate(headers):
            if i < len(row):
                record[key] = row[i]
            else:
                record[key] = ""
        out.append(record)
    return out


def _to_float(val: Any, default: float) -> float:
    """安全轉為 float，失敗時回傳預設值。"""
    try:
        if val in (None, ""):
            return default
        return float(val)
    except (TypeError, ValueError):
        return default


def _to_int(val: Any, default: int) -> int:
    """安全轉為 int，失敗時回傳預設值。"""
    try:
        if val in (None, ""):
            return default
        return int(float(val))
    except (TypeError, ValueError):
        return default


def _to_bool(val: Any, default: bool = False) -> bool:
    if isinstance(val, bool):
        return val
    normalized = str(val or "").strip().casefold()
    if normalized in {"true", "1", "yes", "y"}:
        return True
    if normalized in {"false", "0", "no", "n", ""}:
        return False
    return default


def _validated_non_negative(value: Any, field_name: str) -> float:
    """將營養數值轉成非負浮點數；無效值直接拒絕，避免污染試算表。"""
    return round(finite_non_negative(value, field_name), 2)


def clear_read_caches() -> None:
    """集中清除所有 Sheets 讀取快取；每個寫入操作完成後呼叫。"""
    for fn_name in (
        "get_users_rows",
        "get_records",
        "get_weight_records",
        "get_latest_weight",
        "get_training_records",
        "get_notes",
        "get_audit_logs",
        "get_password_reset_requests",
    ):
        fn = globals().get(fn_name)
        clear = getattr(fn, "clear", None)
        if clear is not None:
            clear()


# =============================================================================
# Users
# =============================================================================

@st.cache_data(ttl=CACHE_TTL)
def get_users_rows() -> list[dict[str, Any]]:
    """取得所有使用者列（教練與學員）。"""
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Users", USERS_HEADERS)
    return _rows_to_dicts(ws, USERS_HEADERS)


def append_user(
    user_id: str,
    username: str,
    name: str,
    password_hash: str,
    created_at: str,
    goals: dict[str, float],
    coach_id: str,
    record_mode: str = "simple",
    weekly_training: int = 4,
) -> None:
    """新增學員到 Users 工作表。

    Args:
        user_id: 學員唯一 ID
        username: 帳號名稱
        password_hash: bcrypt 密碼雜湊
        created_at: 創建時間（ISO 字串）
        goals: 營養目標字典，含 calorie/protein/carb/fat/water
        record_mode: "simple" 或 "full"，飲食記錄模式
        coach_id: 所屬教練 ID
        weekly_training: 每週訓練目標次數，預設 4
    """
    user_id = bounded_text(user_id, "user_id", limit=80)
    username = bounded_text(username, "帳號", limit=TEXT_LIMITS["username"])
    name = bounded_text(name, "姓名", limit=TEXT_LIMITS["name"])
    created_at = valid_timestamp(created_at, "created_at")
    coach_id = bounded_text(coach_id, "coach_id", limit=80)
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Users", USERS_HEADERS)
    row = [
        user_id,
        username,
        name,
        password_hash,
        created_at,
        "",  # BMR 留空，後續由 update_user_bmr 更新
        goals.get("calorie", 0),
        goals.get("protein", 0),
        goals.get("carb", 0),
        goals.get("fat", 0),
        goals.get("water", 0),
        "student",
        weekly_training,
        record_mode,
        coach_id,
        False,
    ]
    ws.append_row(row, value_input_option="USER_ENTERED")
    clear_read_caches()


def _batch_cell(value: Any) -> dict[str, Any]:
    if isinstance(value, bool):
        return {"userEnteredValue": {"boolValue": value}}
    if isinstance(value, (int, float)):
        return {"userEnteredValue": {"numberValue": value}}
    return {"userEnteredValue": {"stringValue": str(value or "")}}


def append_user_with_initial_weight(
    user_id: str,
    username: str,
    name: str,
    password_hash: str,
    created_at: str,
    goals: dict[str, float],
    coach_id: str,
    initial_weight: float,
    record_mode: str = "simple",
    weekly_training: int = 4,
) -> None:
    """Atomically append Users and initial Weight rows in one Sheets batch request."""
    user_id = bounded_text(user_id, "user_id", limit=80)
    username = bounded_text(username, "帳號", limit=TEXT_LIMITS["username"])
    name = bounded_text(name, "姓名", limit=TEXT_LIMITS["name"])
    created_at = valid_timestamp(created_at, "created_at")
    coach_id = bounded_text(coach_id, "coach_id", limit=80)
    weight = positive_number(initial_weight, "weight_kg")
    user_row = [
        user_id,
        username,
        name,
        password_hash,
        created_at,
        "",
        finite_non_negative(goals.get("calorie", 0), "calorie"),
        finite_non_negative(goals.get("protein", 0), "protein"),
        finite_non_negative(goals.get("carb", 0), "carb"),
        finite_non_negative(goals.get("fat", 0), "fat"),
        finite_non_negative(goals.get("water", 0), "water"),
        "student",
        weekly_training,
        record_mode,
        coach_id,
        False,
    ]
    weight_row = [created_at, user_id, round(weight, 1)]
    sh = _get_sheet()
    users_ws = _ensure_worksheet(sh, "Users", USERS_HEADERS)
    weight_ws = _ensure_worksheet(sh, "Weight", WEIGHT_HEADERS)
    _retry_transient(lambda: sh.batch_update(
        {
            "requests": [
                {
                    "appendCells": {
                        "sheetId": users_ws.id,
                        "rows": [{"values": [_batch_cell(value) for value in user_row]}],
                        "fields": "userEnteredValue",
                    }
                },
                {
                    "appendCells": {
                        "sheetId": weight_ws.id,
                        "rows": [{"values": [_batch_cell(value) for value in weight_row]}],
                        "fields": "userEnteredValue",
                    }
                },
            ]
        }
    ))
    clear_read_caches()


def get_user_role(user_id: str) -> str:
    """查詢使用者的 role，回傳 admin、coach 或 student（預設）。"""
    for row in get_users_rows():
        if row.get("user_id") == user_id:
            val = str(row.get("role", "student") or "student")
            return val.strip().lower()
    return "student"


def get_primary_coach_id() -> str:
    """取得並驗證新學員固定歸屬的主教練。"""
    coach = next(
        (
            row
            for row in get_users_rows()
            if str(row.get("user_id", "")).strip() == FIXED_PRIMARY_COACH_ID
        ),
        None,
    )
    if coach is None:
        raise EnvironmentError("固定主教練帳號不存在，請由管理員檢查 Users")

    role = str(coach.get("role", "")).strip().lower()
    if role != "coach":
        raise EnvironmentError("固定主教練帳號必須設定為 role=coach")
    return FIXED_PRIMARY_COACH_ID


def set_user_role(user_id: str, role: str) -> None:
    """將指定 user_id 的 role 設為 admin、coach 或 student。"""
    role = (role or "student").strip().lower()
    if role not in ("student", "coach", "admin"):
        raise ValueError("role 只能是 student、coach 或 admin")
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Users", USERS_HEADERS)
    cell = ws.find(user_id)
    if cell is None:
        raise LookupError("找不到此 user_id")
    role_col = USERS_HEADERS.index("role") + 1
    ws.update_cell(cell.row, role_col, role)
    clear_read_caches()


def get_user_record_mode(user_id: str) -> str:
    """取得使用者的飲食記錄模式，回傳 "simple" 或 "full"。"""
    for row in get_users_rows():
        if row.get("user_id") == user_id:
            mode = str(row.get("record_mode", "simple") or "simple")
            return mode.strip().lower()
    return "simple"


def set_user_record_mode(user_id: str, mode: str) -> None:
    """設定使用者的飲食記錄模式。mode 只能是 "simple" 或 "full"。"""
    mode = (mode or "simple").strip().lower()
    if mode not in ("simple", "full"):
        raise ValueError("record_mode 只能是 simple 或 full")
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Users", USERS_HEADERS)
    cell = ws.find(user_id)
    if cell is None:
        raise LookupError("找不到此 user_id")
    col = USERS_HEADERS.index("record_mode") + 1
    ws.update_cell(cell.row, col, mode)
    clear_read_caches()


def get_user_goals(user_id: str) -> dict[str, float]:
    """取得使用者的每日營養目標。"""
    for row in get_users_rows():
        if row.get("user_id") == user_id:
            return {
                "bmr": _to_float(row.get("bmr"), 0.0),
                "calorie": _to_float(row.get("daily_calorie_goal"), 2000.0),
                "protein": _to_float(row.get("daily_protein_goal"), 60.0),
                "carb": _to_float(row.get("daily_carb_goal"), 250.0),
                "fat": _to_float(row.get("daily_fat_goal"), 65.0),
                "water": _to_float(row.get("daily_water_goal"), 2000.0),
                "weekly_training": float(row.get("weekly_training_goal", 4)),
            }
    return {"bmr": 0.0, "calorie": 2000.0, "protein": 60.0, "carb": 250.0, "fat": 65.0, "water": 2000.0, "weekly_training": 4.0}


def update_user_bmr(user_id: str, bmr: float) -> bool:
    """更新使用者的 BMR 值。成功回傳 True。"""
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Users", USERS_HEADERS)
    rows = get_users_rows()
    bmr_value = _validated_non_negative(bmr, "bmr")
    bmr_col = USERS_HEADERS.index("bmr") + 1
    for idx, row in enumerate(rows, start=2):
        if row.get("user_id") == user_id:
            ws.update_cell(idx, bmr_col, bmr_value)
            clear_read_caches()
            return True
    return False


def update_user_goals(user_id: str, goals: dict[str, float]) -> bool:
    """更新使用者的營養目標（不含 BMR）。成功回傳 True。"""
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Users", USERS_HEADERS)
    rows = get_users_rows()

    field_headers = {
        "calorie": "daily_calorie_goal",
        "protein": "daily_protein_goal",
        "carb": "daily_carb_goal",
        "fat": "daily_fat_goal",
        "water": "daily_water_goal",
    }

    found = False
    for idx, row in enumerate(rows, start=2):
        if row.get("user_id") == user_id:
            found = True
            updates = []
            for key, header in field_headers.items():
                if key in goals:
                    col = USERS_HEADERS.index(header) + 1
                    value = _validated_non_negative(goals[key], key)
                    updates.append(gspread.Cell(idx, col, value))
            if updates:
                ws.update_cells(updates, value_input_option="USER_ENTERED")
                clear_read_caches()
            break

    return found


def get_students_for_coach(coach_id: str) -> list[dict[str, Any]]:
    """取得指定教練所屬學員；未分配學員不會出現在清單中。"""
    return [
        row
        for row in get_users_rows()
        if row.get("role", "").strip().lower() == "student"
        and row.get("coach_id", "").strip() == coach_id
    ]


def get_students_for_manager(manager_id: str) -> list[dict[str, Any]]:
    """管理員可查看全部學員；教練只能查看歸屬自己的學員。"""
    role = get_user_role(manager_id)
    if role == "admin":
        return [row for row in get_users_rows() if row.get("role", "").strip().lower() == "student"]
    if role == "coach":
        return get_students_for_coach(manager_id)
    return []


def get_student_for_coach(user_id: str, coach_id: str) -> dict[str, Any] | None:
    """取得屬於指定教練的學員，不符合歸屬時回傳 None。"""
    for row in get_students_for_coach(coach_id):
        if row.get("user_id") == user_id:
            return row
    return None


def get_student_for_manager(user_id: str, manager_id: str) -> dict[str, Any] | None:
    """依管理者角色及歸屬取得學員。"""
    for row in get_students_for_manager(manager_id):
        if row.get("user_id") == user_id:
            return row
    return None


def user_must_change_password(user_id: str) -> bool:
    """Return whether a user must replace a temporary password."""
    target = str(user_id or "").strip()
    for row in get_users_rows():
        if str(row.get("user_id") or "").strip() == target:
            return _to_bool(row.get("must_change_password"))
    return False


def update_user_password(
    user_id: str, password_hash: str, *, must_change_password: bool
) -> bool:
    """Update only password-related fields for one user."""
    target = bounded_text(user_id, "user_id", limit=80)
    password_hash = bounded_text(password_hash, "password_hash", limit=200)
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Users", USERS_HEADERS)
    for row_idx, row in enumerate(get_users_rows(), start=2):
        if str(row.get("user_id") or "").strip() != target:
            continue
        cells = [
            gspread.Cell(
                row_idx,
                USERS_HEADERS.index("password_hash") + 1,
                password_hash,
            ),
            gspread.Cell(
                row_idx,
                USERS_HEADERS.index("must_change_password") + 1,
                bool(must_change_password),
            ),
        ]
        ws.update_cells(cells, value_input_option="RAW")
        clear_read_caches()
        return True
    return False


# =============================================================================
# PasswordResetRequests（忘記密碼申請）
# =============================================================================

@st.cache_data(ttl=CACHE_TTL)
def get_password_reset_requests(status: str | None = None) -> list[dict[str, Any]]:
    sh = _get_sheet()
    ws = _ensure_worksheet(
        sh, "PasswordResetRequests", PASSWORD_RESET_HEADERS
    )
    rows = _rows_to_dicts(ws, PASSWORD_RESET_HEADERS)
    if status is None:
        return rows
    target = str(status).strip().casefold()
    return [
        row
        for row in rows
        if str(row.get("status") or "").strip().casefold() == target
    ]


def create_password_reset_request(
    request_id: str, user_id: str, requested_at: str
) -> bool:
    """Create one pending request unless the student already has one."""
    request_id = bounded_text(request_id, "request_id", limit=80)
    user_id = bounded_text(user_id, "user_id", limit=80)
    requested_at = valid_timestamp(requested_at, "requested_at")
    if any(
        str(row.get("user_id") or "").strip() == user_id
        for row in get_password_reset_requests("pending")
    ):
        return False
    sh = _get_sheet()
    ws = _ensure_worksheet(
        sh, "PasswordResetRequests", PASSWORD_RESET_HEADERS
    )
    ws.append_row(
        [request_id, user_id, requested_at, "pending", "", ""],
        value_input_option="RAW",
    )
    clear_read_caches()
    return True


def resolve_password_reset_request(
    request_id: str,
    *,
    status: str,
    resolved_at: str,
    resolved_by: str,
) -> bool:
    """Mark a pending password request approved or rejected."""
    request_id = bounded_text(request_id, "request_id", limit=80)
    status = str(status or "").strip().casefold()
    if status not in {"approved", "rejected"}:
        raise ValueError("status 必須是 approved 或 rejected")
    resolved_at = valid_timestamp(resolved_at, "resolved_at")
    resolved_by = bounded_text(resolved_by, "resolved_by", limit=80)
    sh = _get_sheet()
    ws = _ensure_worksheet(
        sh, "PasswordResetRequests", PASSWORD_RESET_HEADERS
    )
    rows = _rows_to_dicts(ws, PASSWORD_RESET_HEADERS)
    for row_idx, row in enumerate(rows, start=2):
        if str(row.get("request_id") or "").strip() != request_id:
            continue
        if str(row.get("status") or "").strip().casefold() != "pending":
            return False
        cells = [
            gspread.Cell(
                row_idx, PASSWORD_RESET_HEADERS.index("status") + 1, status
            ),
            gspread.Cell(
                row_idx,
                PASSWORD_RESET_HEADERS.index("resolved_at") + 1,
                resolved_at,
            ),
            gspread.Cell(
                row_idx,
                PASSWORD_RESET_HEADERS.index("resolved_by") + 1,
                resolved_by,
            ),
        ]
        ws.update_cells(cells, value_input_option="RAW")
        clear_read_caches()
        return True
    return False


def approve_password_reset_request(
    request_id: str,
    user_id: str,
    password_hash: str,
    *,
    resolved_at: str,
    resolved_by: str,
) -> bool:
    """Atomically install a temporary password and approve its pending request."""
    request_id = bounded_text(request_id, "request_id", limit=80)
    user_id = bounded_text(user_id, "user_id", limit=80)
    password_hash = bounded_text(password_hash, "password_hash", limit=200)
    resolved_at = valid_timestamp(resolved_at, "resolved_at")
    resolved_by = bounded_text(resolved_by, "resolved_by", limit=80)
    sh = _get_sheet()
    users_ws = _ensure_worksheet(sh, "Users", USERS_HEADERS)
    reset_ws = _ensure_worksheet(
        sh, "PasswordResetRequests", PASSWORD_RESET_HEADERS
    )
    user_row_idx = next(
        (
            index
            for index, row in enumerate(get_users_rows(), start=2)
            if str(row.get("user_id") or "").strip() == user_id
        ),
        None,
    )
    request_row_idx = next(
        (
            index
            for index, row in enumerate(
                _rows_to_dicts(reset_ws, PASSWORD_RESET_HEADERS), start=2
            )
            if str(row.get("request_id") or "").strip() == request_id
            and str(row.get("user_id") or "").strip() == user_id
            and str(row.get("status") or "").strip().casefold() == "pending"
        ),
        None,
    )
    if user_row_idx is None or request_row_idx is None:
        return False

    def update_cell_request(
        sheet_id: int, row_idx: int, col_idx: int, value: Any
    ) -> dict[str, Any]:
        return {
            "updateCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_idx - 1,
                    "endRowIndex": row_idx,
                    "startColumnIndex": col_idx - 1,
                    "endColumnIndex": col_idx,
                },
                "rows": [{"values": [_batch_cell(value)]}],
                "fields": "userEnteredValue",
            }
        }

    requests = [
        update_cell_request(
            users_ws.id,
            user_row_idx,
            USERS_HEADERS.index("password_hash") + 1,
            password_hash,
        ),
        update_cell_request(
            users_ws.id,
            user_row_idx,
            USERS_HEADERS.index("must_change_password") + 1,
            True,
        ),
        update_cell_request(
            reset_ws.id,
            request_row_idx,
            PASSWORD_RESET_HEADERS.index("status") + 1,
            "approved",
        ),
        update_cell_request(
            reset_ws.id,
            request_row_idx,
            PASSWORD_RESET_HEADERS.index("resolved_at") + 1,
            resolved_at,
        ),
        update_cell_request(
            reset_ws.id,
            request_row_idx,
            PASSWORD_RESET_HEADERS.index("resolved_by") + 1,
            resolved_by,
        ),
    ]
    _retry_transient(lambda: sh.batch_update({"requests": requests}))
    clear_read_caches()
    return True


# =============================================================================
# Records（飲食記錄）
# =============================================================================

def append_record(
    timestamp: str,
    user_id: str,
    meal_type: str,
    food_summary: str,
    calories: float,
    protein: float,
    carb: float,
    fat: float,
    water_ml: float,
    image_url: str,
    portion: float,
) -> None:
    """新增一筆飲食記錄到 Records 工作表。"""
    timestamp = valid_timestamp(timestamp)
    user_id = bounded_text(user_id, "user_id", limit=80)
    meal_type = validate_meal_type(meal_type)
    food_summary = bounded_text(
        food_summary, "食物摘要", limit=TEXT_LIMITS["food_summary"]
    )
    calories = finite_non_negative(calories, "calories")
    protein = finite_non_negative(protein, "protein")
    carb = finite_non_negative(carb, "carb")
    fat = finite_non_negative(fat, "fat")
    water_ml = finite_non_negative(water_ml, "water_ml")
    portion = positive_number(portion, "portion")
    image_url = ""
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Records", RECORDS_HEADERS)
    row = [
        timestamp,
        user_id,
        meal_type,
        food_summary,
        round(calories, 2),
        round(protein, 2),
        round(carb, 2),
        round(fat, 2),
        round(water_ml, 2),
        image_url or "",
        round(portion, 2),
    ]
    ws.append_row(row, value_input_option="USER_ENTERED")
    clear_read_caches()


@st.cache_data(ttl=CACHE_TTL)
def get_records(user_id: str | None = None) -> list[dict[str, Any]]:
    """取得飲食記錄，可依 user_id 篩選。"""
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Records", RECORDS_HEADERS)
    raw = _rows_to_dicts(ws, RECORDS_HEADERS)
    out: list[dict[str, Any]] = []
    for r in raw:
        if user_id is not None and r.get("user_id") != user_id:
            continue
        r["calories"] = _to_float(r.get("calories"), 0.0)
        r["protein"] = _to_float(r.get("protein"), 0.0)
        r["carb"] = _to_float(r.get("carb"), 0.0)
        r["fat"] = _to_float(r.get("fat"), 0.0)
        r["water_ml"] = _to_float(r.get("water_ml"), 0.0)
        r["portion"] = _to_float(r.get("portion"), 1.0)
        out.append(r)
    return out


def get_records_by_date(user_id: str, target_date: date) -> list[dict[str, Any]]:
    """取得指定日期的飲食記錄。target_date 會比對 timestamp 的日期部分（前 10 字）。"""
    date_str = target_date.isoformat()[:10]
    all_records = get_records(user_id)
    return [r for r in all_records if r.get("timestamp", "")[:10] == date_str]


def update_record(record_timestamp: str, user_id: str, updates: dict[str, Any]) -> bool:
    """更新指定飲食記錄的欄位。成功回傳 True。"""
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Records", RECORDS_HEADERS)
    raw_records = ws.get_all_values()

    field_map = {
        "food_summary": 4,
        "calories": 5,
        "protein": 6,
        "carb": 7,
        "fat": 8,
        "water_ml": 9,
        "portion": 11,
    }

    for row_idx, raw_row in enumerate(raw_records[1:], start=2):
        if raw_row and raw_row[0] == record_timestamp and raw_row[1] == user_id:
            for key, col in field_map.items():
                if key in updates:
                    val = updates[key]
                    if isinstance(val, float):
                        val = round(val, 2)
                    ws.update_cell(row_idx, col, val)
            clear_read_caches()
            return True
    return False


def delete_record(record_timestamp: str, user_id: str) -> bool:
    """刪除指定飲食記錄。成功回傳 True。"""
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Records", RECORDS_HEADERS)
    raw_records = ws.get_all_values()

    for row_idx, raw_row in enumerate(raw_records[1:], start=2):
        if raw_row and raw_row[0] == record_timestamp and raw_row[1] == user_id:
            ws.delete_rows(row_idx)
            clear_read_caches()
            return True
    return False


# =============================================================================
# Weight（體重記錄）
# =============================================================================

def append_weight(timestamp: str, user_id: str, weight_kg: float) -> None:
    """新增一筆體重記錄到 Weight 工作表。"""
    timestamp = valid_timestamp(timestamp)
    user_id = bounded_text(user_id, "user_id", limit=80)
    weight_kg = positive_number(weight_kg, "weight_kg")
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Weight", WEIGHT_HEADERS)
    row = [
        timestamp,
        user_id,
        round(weight_kg, 1),
    ]
    ws.append_row(row, value_input_option="USER_ENTERED")
    clear_read_caches()


@st.cache_data(ttl=CACHE_TTL)
def get_weight_records(user_id: str | None = None) -> list[dict[str, Any]]:
    """取得體重記錄，可依 user_id 篩選。"""
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Weight", WEIGHT_HEADERS)
    raw = _rows_to_dicts(ws, WEIGHT_HEADERS)
    out: list[dict[str, Any]] = []
    for r in raw:
        if user_id is not None and r.get("user_id") != user_id:
            continue
        r["weight_kg"] = _to_float(r.get("weight_kg"), 0.0)
        out.append(r)
    return out


@st.cache_data(ttl=CACHE_TTL)
def get_latest_weight(user_id: str) -> float | None:
    """取得學員最新一筆體重記錄，回傳 kg 值，無記錄時回傳 None。"""
    summary = summarize_weight_measurements(get_weight_records(user_id))
    return summary.latest_weight if summary is not None else None


def get_weight_by_date(user_id: str, target_date: date) -> float | None:
    """取得指定日期的體重記錄，無記錄時回傳 None。"""
    date_str = target_date.isoformat()[:10]
    records = [
        record
        for record in get_weight_records(user_id)
        if str(record.get("timestamp", ""))[:10] == date_str
    ]
    summary = summarize_weight_measurements(records)
    return summary.latest_weight if summary is not None else None


# =============================================================================
# Training（訓練記錄）
# =============================================================================

def normalize_training_types(value: str | Iterable[str] | None) -> list[str]:
    """將工作表或表單中的訓練類型正規化為固定順序的清單。"""
    if value is None:
        candidates: list[str] = []
    elif isinstance(value, str):
        candidates = [item.strip() for item in value.split("、") if item.strip()]
    else:
        candidates = [str(item).strip() for item in value if str(item).strip()]

    unknown = set(candidates) - set(TRAINING_TYPE_FIELDS)
    if unknown:
        raise ValueError(f"不支援的訓練類型：{'、'.join(sorted(unknown))}")
    return [label for label in TRAINING_TYPE_FIELDS if label in candidates]


def _validated_training_values(
    training_types: str | Iterable[str] | None,
    strength_detail: str = "",
    cardio_detail: str = "",
    other_detail: str = "",
) -> tuple[list[str], dict[str, str]]:
    selected = normalize_training_types(training_types)
    if not selected:
        raise ValueError("請至少選擇一種訓練類型")

    details = {
        "strength_detail": bounded_text(
            strength_detail, "重量訓練內容", limit=TEXT_LIMITS["training_detail"], required=False
        ),
        "cardio_detail": bounded_text(
            cardio_detail, "有氧訓練內容", limit=TEXT_LIMITS["training_detail"], required=False
        ),
        "other_detail": bounded_text(
            other_detail, "其他訓練內容", limit=TEXT_LIMITS["training_detail"], required=False
        ),
    }
    missing = [label for label in selected if not details[TRAINING_TYPE_FIELDS[label]]]
    if missing:
        raise ValueError(f"請填寫{'、'.join(missing)}內容")
    for label, field in TRAINING_TYPE_FIELDS.items():
        if label not in selected:
            details[field] = ""
    return selected, details


def format_training_record(record: dict[str, Any]) -> str:
    """將一筆訓練紀錄轉成教練頁與匯出共用的人類可讀文字。"""
    types = normalize_training_types(record.get("training_types"))
    parts = []
    for label in types:
        detail = str(record.get(TRAINING_TYPE_FIELDS[label], "") or "").strip()
        parts.append(f"{label}：{detail}" if detail else label)
    return "；".join(parts)

def append_training(
    timestamp: str,
    user_id: str,
    training_types: str | Iterable[str],
    strength_detail: str = "",
    cardio_detail: str = "",
    other_detail: str = "",
) -> None:
    """新增一筆訓練記錄到 Training 工作表。"""
    timestamp = valid_timestamp(timestamp)
    user_id = bounded_text(user_id, "user_id", limit=80)
    selected, details = _validated_training_values(
        training_types, strength_detail, cardio_detail, other_detail
    )
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Training", TRAINING_HEADERS)
    row = [
        timestamp,
        user_id,
        "、".join(selected),
        details["strength_detail"],
        details["cardio_detail"],
        details["other_detail"],
    ]
    ws.append_row(row, value_input_option="USER_ENTERED")
    clear_read_caches()


@st.cache_data(ttl=CACHE_TTL)
def get_training_records(user_id: str | None = None) -> list[dict[str, Any]]:
    """取得訓練記錄，可依 user_id 篩選。"""
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Training", TRAINING_HEADERS)
    raw = _rows_to_dicts(ws, TRAINING_HEADERS)
    out: list[dict[str, Any]] = []
    for r in raw:
        if user_id is not None and r.get("user_id") != user_id:
            continue
        r["training_types"] = normalize_training_types(r.get("training_types"))
        for field in TRAINING_TYPE_FIELDS.values():
            r[field] = str(r.get(field, "") or "").strip()
        out.append(r)
    return out


def get_training_by_date(user_id: str, target_date: date) -> dict[str, Any] | None:
    """取得指定日期的訓練記錄，回傳 dict（無記錄回傳 None）。"""
    date_str = target_date.isoformat()[:10]
    for r in get_training_records(user_id):
        if r.get("timestamp", "")[:10] == date_str:
            return {
                "training_types": r.get("training_types", []),
                "strength_detail": r.get("strength_detail", ""),
                "cardio_detail": r.get("cardio_detail", ""),
                "other_detail": r.get("other_detail", ""),
            }
    return None


def has_training_today(user_id: str, target_date: date) -> bool:
    """檢查指定日期是否有任何訓練記錄。"""
    record = get_training_by_date(user_id, target_date)
    if record is None:
        return False
    return bool(record.get("training_types"))


def update_training(
    timestamp: str,
    user_id: str,
    training_types: str | Iterable[str],
    strength_detail: str = "",
    cardio_detail: str = "",
    other_detail: str = "",
) -> bool:
    """更新指定日期的訓練記錄，若該日記錄不存在則新增。成功回傳 True。"""
    timestamp = valid_timestamp(timestamp)
    user_id = bounded_text(user_id, "user_id", limit=80)
    selected, details = _validated_training_values(
        training_types, strength_detail, cardio_detail, other_detail
    )
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Training", TRAINING_HEADERS)
    raw_records = ws.get_all_values()

    # 先找看看有沒有該日期的記錄
    for row_idx, raw_row in enumerate(raw_records[1:], start=2):
        if raw_row and raw_row[0] == timestamp and raw_row[1] == user_id:
            cells = [
                gspread.Cell(row_idx, col, val)
                for col, val in enumerate(
                    [
                        "、".join(selected),
                        details["strength_detail"],
                        details["cardio_detail"],
                        details["other_detail"],
                    ],
                    start=3,
                )
            ]
            ws.update_cells(cells, value_input_option="USER_ENTERED")
            clear_read_caches()
            return True

    # 沒找到就新增
    append_training(
        timestamp,
        user_id,
        selected,
        details["strength_detail"],
        details["cardio_detail"],
        details["other_detail"],
    )
    return True


# =============================================================================
# Notes（教練備註）
# =============================================================================

def append_note(timestamp: str, user_id: str, coach_id: str, note: str) -> None:
    """新增一筆教練備註到 Notes 工作表。"""
    timestamp = valid_timestamp(timestamp)
    user_id = bounded_text(user_id, "user_id", limit=80)
    coach_id = bounded_text(coach_id, "coach_id", limit=80)
    note = bounded_text(note, "備註", limit=TEXT_LIMITS["note"])
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Notes", NOTES_HEADERS)
    row = [
        timestamp,
        user_id,
        coach_id,
        note,
    ]
    ws.append_row(row, value_input_option="USER_ENTERED")
    clear_read_caches()


@st.cache_data(ttl=CACHE_TTL)
def get_notes(user_id: str | None = None) -> list[dict[str, Any]]:
    """取得教練備註，可依 user_id 篩選。"""
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Notes", NOTES_HEADERS)
    raw = _rows_to_dicts(ws, NOTES_HEADERS)
    out: list[dict[str, Any]] = []
    for r in raw:
        if user_id is not None and r.get("user_id") != user_id:
            continue
        out.append(r)
    return out


def get_latest_note(user_id: str) -> dict[str, Any] | None:
    """取得指定學員的最新一筆備註。"""
    notes = get_notes(user_id)
    if not notes:
        return None
    return notes[-1]  # 因為是時間排序，最後一筆是最新的
def update_note(timestamp: str, user_id: str, coach_id: str, new_note: str) -> bool:
    """更新指定教練備註的內容。成功回傳 True。"""
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Notes", NOTES_HEADERS)
    raw_records = ws.get_all_values()
    for row_idx, raw_row in enumerate(raw_records[1:], start=2):
        if raw_row and raw_row[0] == timestamp and raw_row[1] == user_id and raw_row[2] == coach_id:
            ws.update_cell(row_idx, 4, new_note)  # note 在第 4 欄
            clear_read_caches()
            return True
    return False


def delete_note(timestamp: str, user_id: str, coach_id: str) -> bool:
    """刪除指定教練備註。成功回傳 True。"""
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "Notes", NOTES_HEADERS)
    raw_records = ws.get_all_values()
    for row_idx, raw_row in enumerate(raw_records[1:], start=2):
        if raw_row and raw_row[0] == timestamp and raw_row[1] == user_id and raw_row[2] == coach_id:
            ws.delete_rows(row_idx)
            clear_read_caches()
            return True
    return False


# =============================================================================
# AuditLog（僅保存操作中繼資料，不保存健康內容或 secrets）
# =============================================================================

def append_audit_log(
    *,
    timestamp: str,
    request_id: str,
    actor_id: str,
    actor_role: str,
    action: str,
    target_type: str = "",
    target_id: str = "",
    result: str,
    metadata: dict[str, Any] | None = None,
) -> None:
    timestamp = valid_timestamp(timestamp)
    request_id = bounded_text(request_id, "request_id", limit=80)
    actor_id = bounded_text(actor_id, "actor_id", limit=80, required=False)
    actor_role = bounded_text(actor_role, "actor_role", limit=20, required=False)
    action = bounded_text(action, "action", limit=100)
    target_type = bounded_text(target_type, "target_type", limit=50, required=False)
    target_id = bounded_text(target_id, "target_id", limit=80, required=False)
    result = bounded_text(result, "result", limit=30)
    safe_metadata = {
        str(key)[:60]: value
        for key, value in (metadata or {}).items()
        if str(key).lower() not in {"password", "api_key", "secret", "note", "photo"}
    }
    metadata_json = json.dumps(safe_metadata, ensure_ascii=False, default=str)[:2000]
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "AuditLog", AUDIT_HEADERS)
    ws.append_row(
        [
            timestamp,
            request_id,
            actor_id,
            actor_role,
            action,
            target_type,
            target_id,
            result,
            metadata_json,
        ],
        value_input_option="RAW",
    )


@st.cache_data(ttl=CACHE_TTL)
def get_audit_logs(limit: int = 100) -> list[dict[str, Any]]:
    sh = _get_sheet()
    ws = _ensure_worksheet(sh, "AuditLog", AUDIT_HEADERS)
    rows = _rows_to_dicts(ws, AUDIT_HEADERS)
    return rows[-max(1, min(int(limit), 1000)) :]


def import_records_from_excel(
    excel_file_bytes: bytes = None,
    user_id: str = None,
    overwrite_duplicates: bool = False,
    precomputed_data: dict = None,
    operation_token: str | None = None,
) -> dict[str, Any]:
    """從 Excel 檔案匯入飲食記錄。
    
    固定解析第 8 列（A8:G8）為欄位列。
    只認精確欄位名稱：日期、總熱量(kcal)、蛋白質(g)、喝水(ml)
    
    參數：
    - excel_file_bytes: Excel 檔案位元組（可省略，若有 precomputed_data）
    - user_id: 學員 ID（可省略，若有 precomputed_data）
    - overwrite_duplicates: 是否覆寫已存在的日期記錄
    - precomputed_data: 預計算資料（第一次分析後的結果），用於第二次執行時避免重複讀取
    
    返回：
        imported: 新增數量, 
        overwritten: 覆寫數量, 
        skipped: 跳過數量,
        duplicates: [{date: 日期, existing: 既有資料}],
        errors: [錯誤訊息],
        parsed_data: {records: [...], existing_dates: {...}}（第一次執行時附加）
    }
    """
    import openpyxl
    from io import BytesIO
    from datetime import datetime
    import zipfile
    
    result = {
        "imported": 0,
        "overwritten": 0,
        "skipped": 0,
        "duplicates": [],
        "errors": [],
        "parsed_data": None
    }
    
    # 解析過的記錄（用於第二次執行時直接使用）
    parsed_records = []
    existing_dates = {}
    
    try:
        # 如果有預計算資料，直接使用
        if precomputed_data is not None:
            expected_token = str(precomputed_data.get("operation_token") or "")
            if not operation_token or operation_token != expected_token:
                raise ValueError("匯入操作已失效，請重新分析檔案")
            if str(precomputed_data.get("user_id") or "") != str(user_id or ""):
                raise ValueError("匯入目標學員不一致")
            with _IMPORT_TOKEN_LOCK:
                if operation_token in _APPLIED_IMPORT_TOKENS:
                    raise ValueError("此匯入操作已執行")
            parsed_records = precomputed_data["records"]
            existing_dates = precomputed_data["existing_dates"]
        else:
            # 第一次執行：讀取 Excel 並解析
            if excel_file_bytes is None or user_id is None:
                raise ValueError("excel_file_bytes 和 user_id 是必要參數")
            if len(excel_file_bytes) > SETTINGS.max_upload_mb * 1024 * 1024:
                raise ValueError(f"Excel 檔案不可超過 {SETTINGS.max_upload_mb} MB")
            try:
                with zipfile.ZipFile(BytesIO(excel_file_bytes)) as archive:
                    members = archive.infolist()
                    total_uncompressed = sum(item.file_size for item in members)
                    total_compressed = max(sum(item.compress_size for item in members), 1)
                    if len(members) > 500 or total_uncompressed > 50 * 1024 * 1024:
                        raise ValueError("Excel 解壓縮內容過大")
                    if total_uncompressed / total_compressed > 100:
                        raise ValueError("Excel 壓縮比例異常")
            except zipfile.BadZipFile as exc:
                raise ValueError("Excel 檔案格式無效") from exc

            wb = openpyxl.load_workbook(
                BytesIO(excel_file_bytes), read_only=True, data_only=False
            )
            if len(wb.sheetnames) > 24:
                raise ValueError("Excel 工作表不可超過 24 個")
            
            # 取得現有記錄以檢查重複
            existing_records = get_records(user_id)
            for r in existing_records:
                ts = r.get("timestamp", "")
                if ts:
                    existing_dates[ts[:10]] = r
            
            # 固定欄位名稱（精確匹配）
            TARGET_HEADERS = {
                "date": "日期",
                "calories": "總熱量(kcal)",
                "protein": "蛋白質(g)",
                "water": "喝水(ml)"
            }
            
            # 排除關鍵字
            EXCLUDE_KEYWORDS = ["總分", "合計"]
            
            # 需要轉換為 0 的值
            ZERO_VALUES = ["-", "無", "NA", "", None]
            
            def clean_numeric(val):
                if val is None:
                    return 0.0
                val_str = str(val).strip()
                if val_str in ZERO_VALUES:
                    return 0.0
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return 0.0
            
            def parse_date(date_val):
                if isinstance(date_val, datetime):
                    return date_val.strftime("%Y-%m-%d")
                elif isinstance(date_val, str):
                    # 嘗試 %m/%d 格式（如 6/30）
                    try:
                        parsed = datetime.strptime(date_val.strip(), "%m/%d")
                        return f"{datetime.now().year}-{parsed.month:02d}-{parsed.day:02d}"
                    except ValueError:
                        pass
                    # 嘗試其他常見格式
                    for fmt in ["%Y/%m/%d", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"]:
                        try:
                            return datetime.strptime(date_val.strip(), fmt).strftime("%Y-%m-%d")
                        except ValueError:
                            continue
                return None
            
            for sheet_name in wb.sheetnames:
                # 只處理名稱包含「記錄」的工作表
                if "記錄" not in sheet_name:
                    continue
                    
                ws = wb[sheet_name]
                if ws.max_row > 10008:
                    result["errors"].append(f"工作表「{sheet_name}」超過 10,000 筆資料，略過")
                    continue
                
                # 第 8 列為欄位列
                header_row = 8
                if ws.max_row < header_row:
                    result["errors"].append(f"工作表「{sheet_name}」列數不足，略過")
                    continue
                
                # 讀取 A8:G8
                headers = []
                for col in range(1, 8):
                    cell_val = ws.cell(row=header_row, column=col).value
                    headers.append(cell_val)
                
                # 找出各欄位索引
                col_idx = {key: -1 for key in TARGET_HEADERS}
                for idx, header_val in enumerate(headers):
                    header_str = str(header_val).strip() if header_val else ""
                    for key, target in TARGET_HEADERS.items():
                        if header_str == target:
                            col_idx[key] = idx
                            break
                
                # 檢查必要欄位
                if col_idx["date"] == -1:
                    result["errors"].append(f"工作表「{sheet_name}」找不到「日期」欄位，略過")
                    continue
                if col_idx["calories"] == -1 and col_idx["protein"] == -1:
                    result["errors"].append(f"工作表「{sheet_name}」找不到熱量或蛋白質欄位，略過")
                    continue
                
                # 從第 9 列開始讀取資料
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    try:
                        row = [ws.cell(row=row_idx, column=col).value for col in range(1, 8)]
                        if any(isinstance(value, str) and value.startswith("=") for value in row):
                            result["errors"].append(
                                f"工作表「{sheet_name}」第 {row_idx} 列包含公式，略過"
                            )
                            continue
                        
                        # 取得日期值
                        date_val = row[col_idx["date"]]
                        if date_val is None:
                            continue
                        
                        # 檢查是否為排除關鍵字（如「總分」）
                        date_str = str(date_val).strip()
                        if any(kw in date_str for kw in EXCLUDE_KEYWORDS):
                            continue
                        
                        # 解析日期
                        parsed_date = parse_date(date_val)
                        if parsed_date is None:
                            result["errors"].append(f"工作表「{sheet_name}」第 {row_idx} 列：日期格式無法解析「{date_val}」")
                            continue
                        
                        # 解析數值欄位
                        protein = 0.0
                        calories = 0.0
                        water = 0.0
                        
                        if col_idx["protein"] != -1 and col_idx["protein"] < len(row):
                            protein = clean_numeric(row[col_idx["protein"]])
                        
                        if col_idx["calories"] != -1 and col_idx["calories"] < len(row):
                            calories = clean_numeric(row[col_idx["calories"]])
                        
                        if col_idx["water"] != -1 and col_idx["water"] < len(row):
                            water = clean_numeric(row[col_idx["water"]])
                        
                        # 儲存解析後的記錄
                        parsed_records.append({
                            "date": parsed_date,
                            "timestamp": f"{parsed_date}T12:00:00",
                            "protein": protein,
                            "calories": calories,
                            "water": water
                        })
                        
                    except Exception as e:
                        result["errors"].append(f"工作表「{sheet_name}」第 {row_idx} 列：{str(e)}")
        
        # 第一次執行：只分析不寫入，回傳結果並附加 parsed_data
        # 第二次執行（precomputed_data 不為 None）：執行實際寫入
        if precomputed_data is None:
            # 第一次執行：統計分析結果
            for record in parsed_records:
                if record["date"] in existing_dates:
                    result["duplicates"].append({
                        "date": record["date"],
                        "existing": existing_dates[record["date"]]
                    })
                    result["skipped"] += 1
                else:
                    result["imported"] += 1
            
            # 附加預計算資料供第二次使用
            result["parsed_data"] = {
                "records": parsed_records,
                "existing_dates": existing_dates.copy(),
                "user_id": user_id,
                "file_hash": hashlib.sha256(excel_file_bytes).hexdigest(),
                "operation_token": uuid.uuid4().hex,
            }
        else:
            # 第二次執行：以單一 spreadsheet batch request 套用
            sh = _get_sheet()
            ws = _ensure_worksheet(sh, "Records", RECORDS_HEADERS)
            raw_rows = ws.get_all_values()
            food_rows_by_date: dict[str, int] = {}
            for row_index, raw_row in enumerate(raw_rows[1:], start=2):
                if len(raw_row) < 3 or raw_row[1] != user_id:
                    continue
                if raw_row[2] == "食物" and raw_row[0]:
                    food_rows_by_date.setdefault(raw_row[0][:10], row_index)

            requests: list[dict[str, Any]] = []
            for record in parsed_records:
                record_row = [
                    record["timestamp"], user_id, "食物", "由 Excel 匯入",
                    finite_non_negative(record["calories"], "calories"),
                    finite_non_negative(record["protein"], "protein"),
                    0, 0,
                    finite_non_negative(record["water"], "water"),
                    "", 1.0,
                ]
                if record["date"] in existing_dates:
                    if overwrite_duplicates:
                        row_index = food_rows_by_date.get(record["date"])
                        if row_index is None:
                            result["errors"].append(
                                f"{record['date']} 找不到可安全覆寫的食物列"
                            )
                            continue
                        record_row[3] = "由 Excel 匯入（覆寫）"
                        requests.append(
                            {
                                "updateCells": {
                                    "range": {
                                        "sheetId": ws.id,
                                        "startRowIndex": row_index - 1,
                                        "endRowIndex": row_index,
                                        "startColumnIndex": 0,
                                        "endColumnIndex": len(RECORDS_HEADERS),
                                    },
                                    "rows": [{"values": [_batch_cell(value) for value in record_row]}],
                                    "fields": "userEnteredValue",
                                }
                            }
                        )
                        result["overwritten"] += 1
                    else:
                        result["skipped"] += 1
                else:
                    requests.append(
                        {
                            "appendCells": {
                                "sheetId": ws.id,
                                "rows": [{"values": [_batch_cell(value) for value in record_row]}],
                                "fields": "userEnteredValue",
                            }
                        }
                    )
                    existing_dates[record["date"]] = True
                    result["imported"] += 1
            if requests:
                _retry_transient(lambda: sh.batch_update({"requests": requests}))
                clear_read_caches()
            with _IMPORT_TOKEN_LOCK:
                _APPLIED_IMPORT_TOKENS.add(operation_token)
    
    except Exception as e:
        result["errors"].append(f"讀取 Excel 檔案失敗：{str(e)}")
    
    return result

def overwrite_record_by_date(user_id: str, date_str: str, calories: float, protein: float, water: float) -> bool:
    """根據日期覆寫飲食記錄（用於匯入時覆蓋）。"""
    timestamp = f"{date_str}T12:00:00"
    # 先刪除現有記錄
    delete_record(timestamp, user_id)
    # 重新寫入
    append_record(
        timestamp=timestamp,
        user_id=user_id,
        meal_type="食物",
        food_summary="從 Excel 匯入（覆寫）",
        calories=calories,
        protein=protein,
        carb=0,
        fat=0,
        water_ml=water,
        image_url="",
        portion=1.0,
    )
    return True
